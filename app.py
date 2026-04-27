from __future__ import annotations

import io
import json
import subprocess
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import streamlit as st

from risk_dashboard.aps_cache import default_cache_paths as aps_default_cache_paths
from risk_dashboard.aps_cache import load_cached_tables as aps_load_cached_tables
from risk_dashboard.aps_cache import save_cached_tables as aps_save_cached_tables
from risk_dashboard.aps_cache import signature as aps_signature

# Backward-compatible import (in case app.py is deployed before aps_cache.py update)
try:
    from risk_dashboard.aps_cache import load_any_tables as aps_load_any_tables
except Exception:  # pragma: no cover
    def aps_load_any_tables(paths: object) -> dict[str, pd.DataFrame] | None:
        try:
            pkl = getattr(paths, "data_pkl", None)
            if pkl is None or (hasattr(pkl, "exists") and not pkl.exists()):
                return None
            obj = pd.read_pickle(pkl)
            if not isinstance(obj, dict):
                return None
            out: dict[str, pd.DataFrame] = {}
            for k, v in obj.items():
                if isinstance(v, pd.DataFrame):
                    out[str(k)] = v
            return out or None
        except Exception:
            return None
from risk_dashboard.aps_variation import analyze_workbook as analyze_aps_workbook
from risk_dashboard.io import default_data_paths, load_production_actuals
from risk_dashboard.logging_utils import get_logger
from risk_dashboard.production import month_end, month_start, prev_month_range
from risk_dashboard.production import build_views_with_ranges
from risk_dashboard.master_products import filter_sgwan, load_master_table
from risk_dashboard.schema import PROD_COLS, PROD_REQUIRED_COLS


st.set_page_config(page_title="S관 생산실적 대시보드", layout="wide", initial_sidebar_state="collapsed")
st.title("S관 생산실적 대시보드")

paths = default_data_paths("data")
logger = get_logger("streamlit.app", log_file="logs/streamlit_app.log")

st.markdown(
    """
<style>
  .stApp { background: #f7f2ea; }
  /* 상단 헤더 색상(다른 대시보드처럼) */
  header[data-testid="stHeader"] {
    background: #f7f2ea;
    border-bottom: 1px solid rgba(0,0,0,0.06);
  }
  /* 상단 우측 툴바/메뉴 배경도 투명 처리 */
  div[data-testid="stToolbar"] { background: transparent; }
  div[data-testid="stDecoration"] { background: transparent; }
  /* pills/segment buttons 느낌 */
  .stButton > button { border-radius: 999px; padding: 0.28rem 0.9rem; border: 1px solid rgba(0,0,0,0.18); background: #fff; }
  .stButton > button:hover { border-color: rgba(0,0,0,0.35); }
  div[data-testid="stVerticalBlockBorderWrapper"] { background: rgba(255,255,255,0.55); border-radius: 14px; border: 1px solid rgba(0,0,0,0.06); }
  div[data-testid="stVerticalBlockBorderWrapper"] > div { padding: 0.6rem 0.9rem; }
  /* pills 선택 강조(버전별 DOM 차이가 있어 범용 selector로 커버) */
  div[data-testid="stPills"] [aria-selected="true"] { background: rgba(46,125,50,0.16) !important; border-color: rgba(46,125,50,0.55) !important; color: #1b5e20 !important; }
  div[data-testid="stPills"] [aria-selected="false"] { background: #fff !important; }
</style>
""",
    unsafe_allow_html=True,
)

# 기본 사이드바 접기


def _norm_text(x: object) -> str:
    if pd.isna(x):
        return ""
    return " ".join(str(x).strip().split())


def _map_process(code: object) -> str:
    s = _norm_text(code)
    if s.startswith("[10]"):
        return "사출조립"
    if s.startswith("[20]"):
        return "분리"
    if s.startswith("[45]"):
        return "하이드레이션/전면검"
    if s.startswith("[55]"):
        return "접착/멸균"
    if s.startswith("[80]"):
        return "누수/규격검사"
    if s.startswith("[85]"):
        return "포장"
    return s


def _excel_upload_time_kst(path: Path) -> str | None:
    if not path.exists():
        return None
    try:
        import openpyxl  # type: ignore[import-not-found]

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        dt = wb.properties.modified or wb.properties.created
        wb.close()
        if dt is None:
            return None
        if getattr(dt, "tzinfo", None) is None:
            dt = dt.replace(tzinfo=ZoneInfo("Asia/Seoul"))
        else:
            dt = dt.astimezone(ZoneInfo("Asia/Seoul"))
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        # fallback (may reflect deploy/copy time on some environments)
        try:
            dt = datetime.fromtimestamp(path.stat().st_mtime, tz=ZoneInfo("Asia/Seoul"))
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return None


UPDATE_DATA_DIR = Path("업데이트 데이터")


def _pick_input_path(filename: str) -> Path:
    """
    사용자가 갱신해야 하는 입력 엑셀은 기본적으로 `업데이트 데이터/`에 둡니다.
    (기존 루트 경로도 남아있으면 하위호환으로 사용)
    """
    new_path = UPDATE_DATA_DIR / filename
    old_path = Path(filename)
    if new_path.exists():
        return new_path
    if old_path.exists():
        return old_path
    return new_path


@st.cache_data(show_spinner=False)
def _excel_sheet_names(path: str, *, cache_bust: float | None = None) -> list[str]:
    _ = cache_bust
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(str(p))
    import openpyxl  # type: ignore[import-not-found]

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


@st.cache_data(show_spinner=False)
def load_order_status_sgwan(
    order_xlsx: str,
    master_xlsx: str,
    *,
    sheet_name: str = "data",
    header_row_1based: int = 2,
    item_name_col: str = "품명",
    master_name_col: str = "제품명",
    cache_bust: float | None = None,
) -> pd.DataFrame:
    _ = cache_bust
    op = Path(order_xlsx)
    mp = Path(master_xlsx)
    if not op.exists():
        raise FileNotFoundError(str(op))
    if not mp.exists():
        raise FileNotFoundError(str(mp))

    raw = pd.read_excel(op, sheet_name=sheet_name, header=None)  # type: ignore[call-arg]
    if raw.empty:
        return pd.DataFrame()
    header_idx = max(0, header_row_1based - 1)
    if header_idx >= len(raw):
        raise ValueError(f"header_row out of range: {header_row_1based} (rows={len(raw)})")
    header = raw.iloc[header_idx].tolist()
    df = raw.iloc[header_idx + 1 :].copy()
    df.columns = [str(c).strip() for c in header]
    df = df.dropna(axis=1, how="all").copy()

    if item_name_col not in df.columns:
        raise ValueError(f"입력에 품명 컬럼이 없습니다: {item_name_col} (컬럼={list(df.columns)})")

    mdf = load_master_table(mp)
    mdf_s = filter_sgwan(mdf)
    if master_name_col not in mdf_s.columns:
        raise ValueError(f"마스터에 제품명 컬럼이 없습니다: {master_name_col} (컬럼={list(mdf_s.columns)})")

    master_names = {_norm_text(x) for x in mdf_s[master_name_col].tolist() if _norm_text(x)}
    summary_col = "분류요약" if "분류요약" in mdf_s.columns else None
    master_summary_map: dict[str, str] = {}
    if summary_col:
        for _, row in mdf_s[[master_name_col, summary_col]].iterrows():
            n = _norm_text(row.get(master_name_col))
            s = _norm_text(row.get(summary_col))
            if n and n not in master_summary_map:
                master_summary_map[n] = s
    if not master_names:
        return pd.DataFrame()

    df2 = df.copy()
    df2["_품명_norm"] = df2[item_name_col].map(_norm_text)
    out = df2[df2["_품명_norm"].isin(master_names)].copy()
    if master_summary_map:
        out["분류요약"] = out["_품명_norm"].map(master_summary_map).fillna("")
    out = out.drop(columns=["_품명_norm"], errors="ignore")
    return out


@st.cache_data(show_spinner=False)
def load_order_status_raw(
    order_xlsx: str,
    *,
    sheet_name: str = "data",
    header_row_1based: int = 2,
    cache_bust: float | None = None,
) -> pd.DataFrame:
    _ = cache_bust
    op = Path(order_xlsx)
    if not op.exists():
        raise FileNotFoundError(str(op))

    raw = pd.read_excel(op, sheet_name=sheet_name, header=None)  # type: ignore[call-arg]
    if raw.empty:
        return pd.DataFrame()
    header_idx = max(0, header_row_1based - 1)
    if header_idx >= len(raw):
        raise ValueError(f"header_row out of range: {header_row_1based} (rows={len(raw)})")
    header = raw.iloc[header_idx].tolist()
    df = raw.iloc[header_idx + 1 :].copy()
    df.columns = [str(c).strip() for c in header]
    df = df.dropna(axis=1, how="all").copy()
    return df


def _normalize_misc_orders_for_dashboard(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    out = df.copy()

    # Derive month key (prefer due date -> request date)
    due_dt = (
        pd.to_datetime(out["납기일자"], errors="coerce")
        if "납기일자" in out.columns
        else pd.Series([pd.NaT] * len(out), index=out.index)
    )
    req_dt = (
        pd.to_datetime(out["요청일자"], errors="coerce")
        if "요청일자" in out.columns
        else pd.Series([pd.NaT] * len(out), index=out.index)
    )
    base_dt = due_dt
    if base_dt.isna().all():
        base_dt = req_dt
    month_dt = base_dt.dt.to_period("M").dt.to_timestamp()

    out["__month_date__"] = month_dt
    out["연도"] = base_dt.dt.year

    # Dashboard expects these columns for monthly aggregation/view
    out["구분"] = "기타수주"

    # 기타수주는 요청번호(작지번호)가 없으므로 공란 처리
    out["작지번호"] = pd.Series([""] * len(out), index=out.index, dtype="string")

    # Quantities: prefer PCS, else PACK * 입수(낱개)
    qty_pcs = (
        pd.to_numeric(out["수량(PCS)"], errors="coerce")
        if "수량(PCS)" in out.columns
        else pd.Series([pd.NA] * len(out), index=out.index)
    )
    if qty_pcs.isna().all():
        qty_pack = (
            pd.to_numeric(out["수량(PACK)"], errors="coerce")
            if "수량(PACK)" in out.columns
            else pd.Series([pd.NA] * len(out), index=out.index)
        )
        pack_in = (
            pd.to_numeric(out["입수(낱개)"], errors="coerce")
            if "입수(낱개)" in out.columns
            else pd.Series([pd.NA] * len(out), index=out.index)
        )
        qty_pcs = qty_pack * pack_in
    out["오더수량"] = qty_pcs.fillna(0)

    out["수주금액(원)"] = 0
    out["수주금액(달러)"] = 0

    if "상태" in out.columns:
        out["현재상태"] = out["상태"]

    # Align key date columns used in the table (keep originals too)
    if "요청일자" in out.columns:
        out["수주 전송일"] = req_dt
    if "납기일자" in out.columns:
        out["영업출고요청일"] = due_dt

    if "포장완료" in out.columns:
        done = pd.to_numeric(out["포장완료"], errors="coerce").fillna(0)
        out["포장 진도율"] = (done > 0).astype(int) * 100

    # Domestic request: normalize country if possible
    if "국내/해외" in out.columns and "국가" not in out.columns:
        out["국가"] = out["국내/해외"].map(lambda x: "대한민국" if _norm_text(x) == "국내" else _norm_text(x))

    if "고객" not in out.columns:
        out["고객"] = "국내요청"
    if "담당자" not in out.columns and "생성자" in out.columns:
        out["담당자"] = out["생성자"]

    return out


@st.cache_data(show_spinner=False)
def load_prod_from_excel(prod_xlsx: str, *, status: str = "확인", cache_bust: float | None = None) -> pd.DataFrame:
    # cache_bust: 파일 수정시간 등을 넘겨 캐시가 파일 변경을 감지하도록 함.
    _ = cache_bust
    p = Path(prod_xlsx)
    if not p.exists():
        raise FileNotFoundError(str(p))
    xl = pd.ExcelFile(p)
    frames = []
    for sh in ["전월", "당월"]:
        if sh in xl.sheet_names:
            frames.append(pd.read_excel(p, sheet_name=sh))  # type: ignore[call-arg]
    if not frames:
        return pd.DataFrame(columns=PROD_REQUIRED_COLS)

    raw = pd.concat(frames, ignore_index=True)
    # prefer 샘플 제외 양품수량(있으면) else 양품수량
    good_col = "샘플제외 양품수량" if "샘플제외 양품수량" in raw.columns else "양품수량"
    required = ["생산일자", "공정코드", "품목코드", "생산수량", good_col, "상태"]
    missing = [c for c in required if c not in raw.columns]
    if missing:
        raise ValueError(f"필수 컬럼 누락: {missing}")

    raw = raw.copy()
    raw["상태_norm"] = raw["상태"].map(_norm_text)
    raw = raw[raw["상태_norm"] == _norm_text(status)].copy()

    out = pd.DataFrame(
        {
            PROD_COLS.생산일자: pd.to_datetime(raw["생산일자"], errors="coerce").dt.date,
            PROD_COLS.공정: raw["공정코드"].map(_map_process),
            PROD_COLS.품목코드: raw["품목코드"].map(_norm_text),
            PROD_COLS.생산수량: pd.to_numeric(raw["생산수량"], errors="coerce"),
            PROD_COLS.양품수량: pd.to_numeric(raw[good_col], errors="coerce"),
            PROD_COLS.신규분류요약: raw.get("신규분류요약", pd.Series([None] * len(raw))).map(_norm_text),
        }
    )
    out = out.dropna(subset=[PROD_COLS.생산일자, PROD_COLS.생산수량]).copy()
    out = out[out[PROD_COLS.품목코드].astype(str).str.len() > 0].copy()
    out[PROD_COLS.생산수량] = out[PROD_COLS.생산수량].astype(int)
    out[PROD_COLS.양품수량] = pd.to_numeric(out[PROD_COLS.양품수량], errors="coerce").fillna(0).astype(int)
    cols = PROD_REQUIRED_COLS + [PROD_COLS.양품수량]
    if PROD_COLS.신규분류요약 in out.columns:
        cols.append(PROD_COLS.신규분류요약)
    return out[cols].copy()


@st.cache_data(show_spinner=False)
def load_aps_risk_tables(
    aps_xlsx: str,
    *,
    scope_processes: tuple[str, ...],
    cache_bust: float | None = None,
) -> dict[str, pd.DataFrame]:
    _ = cache_bust
    return analyze_aps_workbook(aps_xlsx, scope_processes=scope_processes)


@st.cache_data(show_spinner=False)
def load_orders_sgwan_for_join(
    order_book_xlsx: str,
    master_xlsx: str,
    *,
    order_sheet: str,
    misc_sheet: str | None,
    cache_bust: float | None = None,
) -> pd.DataFrame:
    _ = cache_bust
    orders_s = load_order_status_sgwan(
        order_book_xlsx,
        master_xlsx,
        sheet_name=order_sheet,
        cache_bust=cache_bust,
    )

    misc_s = pd.DataFrame()
    if misc_sheet:
        misc_s = load_order_status_sgwan(
            order_book_xlsx,
            master_xlsx,
            sheet_name=misc_sheet,
            header_row_1based=1,
            cache_bust=cache_bust,
        )
        misc_s = _normalize_misc_orders_for_dashboard(misc_s)

    out = pd.concat([orders_s, misc_s], ignore_index=True, sort=False) if not misc_s.empty else orders_s
    if out.empty:
        return out

    # 제품명코드 매핑(품명 -> 제품명코드)
    mdf = load_master_table(Path(master_xlsx))
    mdf_s = filter_sgwan(mdf)
    if "제품명" in mdf_s.columns and "제품명코드" in mdf_s.columns and "품명" in out.columns:
        name_to_code = dict(
            zip(
                mdf_s["제품명"].astype(str).str.strip().tolist(),
                mdf_s["제품명코드"].astype(str).str.strip().tolist(),
            )
        )
        out["제품명코드"] = out["품명"].astype(str).str.strip().map(name_to_code)

    # 날짜 컬럼 normalize
    for col in ["영업출고요청일", "수주 전송일", "영업협의출고일", "포장완료일"]:
        if col in out.columns:
            out[col] = pd.to_datetime(out[col], errors="coerce").dt.date

    # 요청 기준일(조인/관리 기준): 영업협의출고일 우선, 없으면 영업출고요청일
    if "영업협의출고일" in out.columns or "영업출고요청일" in out.columns:
        agreed = pd.to_datetime(out.get("영업협의출고일"), errors="coerce")
        req = pd.to_datetime(out.get("영업출고요청일"), errors="coerce")
        out["요청기준일"] = agreed.where(agreed.notna(), req).dt.date

    # 수량/금액 normalize
    for col in ["오더수량", "수주금액", "수주금액(원)", "수주금액(달러)"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)

    return out


def _aps_status_path(data_dir: str | Path = "data") -> Path:
    return Path(data_dir) / "aps_risk_cache_status.json"


def _aps_lock_path(data_dir: str | Path = "data") -> Path:
    return Path(data_dir) / "aps_risk_cache.lock"


def _read_json_if_exists(path: Path) -> dict | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _launch_aps_cache_build(*, input_path: Path, scope: str, days: int = 8) -> None:
    cmd = [
        sys.executable,
        str(Path("scripts") / "build_aps_risk_cache.py"),
        "--input",
        str(input_path),
        "--scope",
        scope,
        "--days",
        str(days),
    ]
    flags = 0
    if hasattr(subprocess, "CREATE_NO_WINDOW"):
        flags = subprocess.CREATE_NO_WINDOW  # type: ignore[attr-defined]
    subprocess.Popen(  # noqa: S603
        cmd,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        creationflags=flags,
    )


with st.sidebar:
    st.header("데이터")
    default_prod = Path("data/production_actuals_recent.csv")
    default_excel = _pick_input_path("생산실적현황(간편)_S관.xlsx")
    # 원천은 '생산실적현황(간편)_S관.xlsx'를 우선으로 사용(존재하면 기본값을 엑셀로).
    default_source_idx = 1 if default_excel.exists() else 0
    source = st.radio("소스", ["CSV", "엑셀"], index=default_source_idx, horizontal=True)

    if source == "CSV":
        prod_path = st.text_input(
            "생산실적 CSV",
            value=str(default_prod if default_prod.exists() else paths.production_actuals_csv),
        )
        st.caption("CSV 인코딩은 `utf-8-sig` 권장(엑셀 호환).")
    else:
        prod_xlsx = st.text_input("생산실적(간편)", value=str(default_excel))
        st.caption("엑셀 파일이 리포지토리(같은 폴더)에 있으면 바로 동작합니다.")

    with st.expander("수주현황 데이터", expanded=False):
        order_book_xlsx = st.text_input("수주현황 포함 엑셀", value=str(default_excel))
        if st.button("수주현황 시트 새로고침"):
            st.cache_data.clear()
            st.rerun()
        try:
            sheet_names = _excel_sheet_names(
                order_book_xlsx,
                cache_bust=(Path(order_book_xlsx).stat().st_mtime if Path(order_book_xlsx).exists() else None),
            )
        except Exception as e:
            sheet_names = []
            st.error(f"수주현황 엑셀 시트 목록을 읽지 못했습니다: {e}")
        if sheet_names:
            prefer = ["수주현황", "수주 현황", "수주"]
            default_sheet = next((s for s in prefer if s in sheet_names), sheet_names[0])
            order_sheet = st.selectbox("수주현황 시트", options=sheet_names, index=sheet_names.index(default_sheet))

            misc_prefer = ["기타수주현황", "기타 수주현황", "기타수주"]
            misc_sheet = next((s for s in misc_prefer if s in sheet_names), None)
            if misc_sheet is None:
                misc_sheet = next((s for s in sheet_names if "기타수주" in s.replace(" ", "")), None)
            if misc_sheet:
                st.caption(f"기타수주 시트 포함: {misc_sheet}")
        else:
            st.info("엑셀에 '수주현황' 시트를 저장/업로드한 뒤 다시 확인하세요.")
            order_sheet = None
            misc_sheet = None
        master_xlsx = st.text_input("S관 제품 마스터", value=str(_pick_input_path("S관 생산 제품 리스트.xlsx")))

if source == "CSV":
    logger.info("source=CSV | prod=%s", prod_path)
    try:
        prod_df = load_production_actuals(Path(prod_path))
    except Exception as e:
        logger.exception("production load failed")
        st.error(f"생산실적 로드 실패: {e}")
        prod_df = pd.DataFrame()
else:
    logger.info("source=EXCEL | prod=%s", prod_xlsx)
    with st.spinner("엑셀에서 생산실적 생성 중..."):
        try:
            excel_mtime = Path(prod_xlsx).stat().st_mtime if Path(prod_xlsx).exists() else None
            prod_df = load_prod_from_excel(prod_xlsx, status="확인", cache_bust=excel_mtime)
        except Exception as e:
            logger.exception("excel->production failed")
            st.error(f"엑셀 → 생산실적 변환 실패: {e}")
            prod_df = pd.DataFrame(columns=PROD_REQUIRED_COLS)

def _normalize_process_names(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or PROD_COLS.공정 not in df.columns:
        return df
    m = {
        "사출": "사출조립",
        "하이드": "하이드레이션/전면검",
        "접착": "접착/멸균",
        "누수": "누수/규격검사",
    }
    out = df.copy()
    out[PROD_COLS.공정] = out[PROD_COLS.공정].map(lambda x: m.get(str(x), str(x)))
    return out

prod_df = _normalize_process_names(prod_df)

with st.sidebar:
    asof = st.date_input("기준일", value=date.today())

tabs = st.tabs(["S관 실적", "S관 수주현황", "APS 리스크관리"])

with tabs[0]:
    st.subheader("S관 생산실적 현황(간편)")
    # 컨셉 문구는 숨김(요청사항)

    if prod_df.empty:
        st.info("생산실적 데이터가 없습니다.")
    else:
        cutoff = asof - timedelta(days=1)

        # 업로드(원천 엑셀 파일 수정시간 기준) 표시
        excel_path = Path(prod_xlsx) if source == "엑셀" else Path(prod_path)
        uploaded_at = _excel_upload_time_kst(excel_path)
        if uploaded_at:
            st.caption(f"생산실적 업로드 시간 : {uploaded_at} (한국시간)")

        # 전월/당월은 월 단위 고정(당월은 cutoff=전일까지만)
        prev_start, prev_end = prev_month_range(asof)
        curr_start_default = month_start(asof)
        curr_end_default = cutoff

        left_col, right_col = st.columns(2)
        with right_col:
            right_card = st.container(border=True)
        with left_col:
            left_card = st.container(border=True)

        with right_card:
            st.markdown("**당월**")

        curr_start = curr_start_default
        curr_end = curr_end_default

        # 전월/당월 범위 집계
        # - 전월: 월 전체
        # - 당월: 월 시작 ~ cutoff(전일) 고정

        views = build_views_with_ranges(
            prod_df,
            asof=asof,
            prev_start=prev_start,
            prev_end=prev_end,
            curr_start=curr_start,
            curr_end=curr_end,
        )

        def _n_days(df: pd.DataFrame) -> int:
            if df.empty:
                return 0
            return int(df[PROD_COLS.생산일자].nunique())

        def _n_items(df: pd.DataFrame) -> int:
            if df.empty:
                return 0
            return int(df[PROD_COLS.품목코드].nunique())

        process_order = ["사출조립", "분리", "하이드레이션/전면검", "접착/멸균", "누수/규격검사"]

        def _process_summary(df: pd.DataFrame) -> pd.DataFrame:
            if df.empty:
                return pd.DataFrame(columns=[PROD_COLS.공정, "생산수량", "양품수량", "수율"])

            gross_col = PROD_COLS.생산수량
            good_col = PROD_COLS.양품수량 if PROD_COLS.양품수량 in df.columns else gross_col
            g = (
                df.groupby(PROD_COLS.공정, dropna=False)[[gross_col, good_col]]
                .sum()
                .reset_index()
                .rename(columns={gross_col: "생산수량", good_col: "양품수량"})
            )
            g[PROD_COLS.공정] = pd.Categorical(g[PROD_COLS.공정].astype(str), categories=process_order, ordered=True)
            g = g.sort_values(PROD_COLS.공정).copy()
            g["생산수량"] = pd.to_numeric(g["생산수량"], errors="coerce").fillna(0).astype(int)
            g["양품수량"] = pd.to_numeric(g["양품수량"], errors="coerce").fillna(0).astype(int)
            g["수율"] = g.apply(lambda r: (r["양품수량"] / r["생산수량"]) if r["생산수량"] > 0 else None, axis=1)
            return g.reset_index(drop=True)

        def _daily_process_summary(df: pd.DataFrame) -> pd.DataFrame:
            if df.empty:
                return pd.DataFrame(columns=[PROD_COLS.생산일자, PROD_COLS.공정, "생산수량", "양품수량", "수율"])

            gross_col = PROD_COLS.생산수량
            good_col = PROD_COLS.양품수량 if PROD_COLS.양품수량 in df.columns else gross_col

            out = (
                df.groupby([PROD_COLS.생산일자, PROD_COLS.공정], dropna=False)[[gross_col, good_col]]
                .sum()
                .reset_index()
                .rename(columns={gross_col: "생산수량", good_col: "양품수량"})
            )
            out[PROD_COLS.공정] = pd.Categorical(out[PROD_COLS.공정].astype(str), categories=process_order, ordered=True)
            out = out.sort_values([PROD_COLS.생산일자, PROD_COLS.공정]).copy()
            out["생산수량"] = pd.to_numeric(out["생산수량"], errors="coerce").fillna(0).astype(int)
            out["양품수량"] = pd.to_numeric(out["양품수량"], errors="coerce").fillna(0).astype(int)
            out["수율"] = out.apply(lambda r: (r["양품수량"] / r["생산수량"]) if r["생산수량"] > 0 else None, axis=1)
            return out.reset_index(drop=True)

        def _total_output_and_yield(df: pd.DataFrame) -> tuple[int, float | None]:
            if df.empty:
                return 0, None

            gross_col = PROD_COLS.생산수량
            good_col = PROD_COLS.양품수량 if PROD_COLS.양품수량 in df.columns else gross_col
            final_proc = "누수/규격검사"

            final_df = df[df[PROD_COLS.공정].astype(str) == final_proc].copy()
            total_output = int(pd.to_numeric(final_df[good_col], errors="coerce").fillna(0).sum()) if not final_df.empty else 0

            comp = None
            comp_val = 1.0
            for proc in process_order:
                sub = df[df[PROD_COLS.공정].astype(str) == proc].copy()
                gross = int(pd.to_numeric(sub[gross_col], errors="coerce").fillna(0).sum()) if not sub.empty else 0
                good = int(pd.to_numeric(sub[good_col], errors="coerce").fillna(0).sum()) if not sub.empty else 0
                if gross <= 0:
                    comp = None
                    break
                comp_val *= (good / gross)
                comp = float(comp_val)

            return total_output, comp

        def _category_final_and_comp_yield(df: pd.DataFrame, *, asof: date) -> pd.DataFrame:
            if df.empty or PROD_COLS.신규분류요약 not in df.columns:
                return pd.DataFrame(columns=[PROD_COLS.신규분류요약, "생산수량", "양품수량", "종합수율"])

            gross_col = PROD_COLS.생산수량
            good_col = PROD_COLS.양품수량 if PROD_COLS.양품수량 in df.columns else gross_col
            final_proc = "누수/규격검사"

            base = df.copy()
            base[PROD_COLS.신규분류요약] = base[PROD_COLS.신규분류요약].map(_norm_text)
            base = base[base[PROD_COLS.신규분류요약].astype(str).str.len() > 0].copy()
            if base.empty:
                return pd.DataFrame(columns=[PROD_COLS.신규분류요약, "생산수량", "양품수량", "종합수율"])

            # final quantities (누수/규격검사 기준)
            final_df = base[base[PROD_COLS.공정].astype(str) == final_proc].copy()
            final_sum = (
                final_df.groupby(PROD_COLS.신규분류요약, dropna=False)[[gross_col, good_col]]
                .sum()
                .rename(columns={gross_col: "생산수량", good_col: "양품수량"})
            )

            # composite yield per category = product of per-process yields
            proc_sum = (
                base.groupby([PROD_COLS.신규분류요약, PROD_COLS.공정], dropna=False)[[gross_col, good_col]]
                .sum()
                .reset_index()
            )
            proc_sum["수율"] = proc_sum.apply(
                lambda r: (r[good_col] / r[gross_col]) if float(r[gross_col]) > 0 else None,
                axis=1,
            )

            # pivot to ensure all processes exist
            yield_map = proc_sum.pivot_table(
                index=PROD_COLS.신규분류요약,
                columns=PROD_COLS.공정,
                values="수율",
                aggfunc="first",
            )

            comp_vals: dict[str, float] = {}
            for cat, row in yield_map.iterrows():
                ys: list[float] = []
                for proc in process_order:
                    y = row.get(proc)
                    if y is None or pd.isna(y):
                        continue
                    ys.append(float(y))
                if not ys:
                    continue
                prod_val = 1.0
                for y in ys:
                    prod_val *= y
                comp_vals[str(cat)] = float(prod_val)

            comp_series = pd.Series(comp_vals, name="종합수율")

            out = final_sum.join(comp_series, how="left").reset_index()
            out["생산수량"] = pd.to_numeric(out["생산수량"], errors="coerce").fillna(0).astype(int)
            out["양품수량"] = pd.to_numeric(out["양품수량"], errors="coerce").fillna(0).astype(int)
            out = out.sort_values("양품수량", ascending=False).reset_index(drop=True)
            return out[[PROD_COLS.신규분류요약, "생산수량", "양품수량", "종합수율"]].copy()

        def _forecast_eom_output_mtd(df: pd.DataFrame, *, asof: date) -> int | None:
            # Forecast end-of-month output using final-process good qty MTD.
            # - MTD uses cutoff (asof-1) already in views.curr_month.df
            # - average per observed production day (final process) * estimated remaining production days
            if df.empty:
                return None

            gross_col = PROD_COLS.생산수량
            good_col = PROD_COLS.양품수량 if PROD_COLS.양품수량 in df.columns else gross_col
            final_proc = "누수/규격검사"

            cutoff_local = asof - timedelta(days=1)
            ms = month_start(asof)
            me = month_end(asof)
            if cutoff_local < ms:
                return 0

            final_df = df[df[PROD_COLS.공정].astype(str) == final_proc].copy()
            if final_df.empty:
                return 0

            mtd_qty = int(pd.to_numeric(final_df[good_col], errors="coerce").fillna(0).sum())
            prod_days = int(pd.to_datetime(final_df[PROD_COLS.생산일자], errors="coerce").dt.date.nunique())
            if prod_days <= 0:
                return None

            avg_per_day = mtd_qty / prod_days

            elapsed_days = (cutoff_local - ms).days + 1
            if elapsed_days <= 0:
                return None

            # Estimate remaining production days by observed ratio of production-days to elapsed calendar-days.
            prod_day_ratio = min(1.0, max(0.0, prod_days / elapsed_days))
            remaining_days = max(0, (me - cutoff_local).days)
            remaining_prod_days_est = int(round(remaining_days * prod_day_ratio))
            forecast = int(round(mtd_qty + avg_per_day * remaining_prod_days_est))
            return forecast

        with left_card:
            st.markdown("**전월**")
            st.caption(f"생산일수: {_n_days(views.prev_month.df):,} / 품목수: {_n_items(views.prev_month.df):,}")

            st.markdown("**공정별 요약**")
            prev_proc = _process_summary(views.prev_month.df)
            prev_total, prev_comp = _total_output_and_yield(views.prev_month.df)
            k1, k2 = st.columns(2)
            k1.metric("총 생산실적(누수/규격검사 양품)", f"{prev_total:,}")
            k2.metric("종합 수율", f"{prev_comp*100:.1f}%" if prev_comp is not None else "-")
            st.dataframe(
                prev_proc.style.format({"생산수량": "{:,.0f}", "양품수량": "{:,.0f}", "수율": "{:.1%}"}),
                use_container_width=True,
                hide_index=True,
            )

            st.markdown("**신규분류요약별**")
            prev_cat = _category_final_and_comp_yield(views.prev_month.df, asof=asof)
            if prev_cat.empty:
                st.info("신규분류요약 데이터가 없어 그룹 요약을 표시할 수 없습니다.")
            else:
                st.dataframe(
                    prev_cat.style.format({"생산수량": "{:,.0f}", "양품수량": "{:,.0f}", "종합수율": "{:.1%}"}),
                    use_container_width=True,
                    hide_index=True,
                )

            st.markdown("**일자별 집계**")
            prev_daily = _daily_process_summary(views.prev_month.df)
            st.dataframe(
                prev_daily.style.format({"생산수량": "{:,.0f}", "양품수량": "{:,.0f}", "수율": "{:.1%}"}),
                use_container_width=True,
                hide_index=True,
            )

        with right_card:
            st.caption(f"생산일수: {_n_days(views.curr_month.df):,} / 품목수: {_n_items(views.curr_month.df):,}")

            st.markdown("**공정별 요약**")
            curr_proc = _process_summary(views.curr_month.df)
            curr_total, curr_comp = _total_output_and_yield(views.curr_month.df)
            k1, k2, k3 = st.columns(3)
            k1.metric("총 생산실적(누수/규격검사 양품)", f"{curr_total:,}")
            k2.metric("종합 수율", f"{curr_comp*100:.1f}%" if curr_comp is not None else "-")
            curr_forecast = _forecast_eom_output_mtd(views.curr_month.df, asof=asof)
            k3.metric("월말 예상(누수/규격검사 양품)", f"{curr_forecast:,}" if curr_forecast is not None else "-")
            st.dataframe(
                curr_proc.style.format({"생산수량": "{:,.0f}", "양품수량": "{:,.0f}", "수율": "{:.1%}"}),
                use_container_width=True,
                hide_index=True,
            )

            st.markdown("**신규분류요약별**")
            curr_cat = _category_final_and_comp_yield(views.curr_month.df, asof=asof)
            if curr_cat.empty:
                st.info("신규분류요약 데이터가 없어 그룹 요약을 표시할 수 없습니다.")
            else:
                st.dataframe(
                    curr_cat.style.format({"생산수량": "{:,.0f}", "양품수량": "{:,.0f}", "종합수율": "{:.1%}"}),
                    use_container_width=True,
                    hide_index=True,
                )

            st.markdown("**일자별 집계**")
            curr_daily = _daily_process_summary(views.curr_month.df)
            st.dataframe(
                curr_daily.style.format({"생산수량": "{:,.0f}", "양품수량": "{:,.0f}", "수율": "{:.1%}"}),
                use_container_width=True,
                hide_index=True,
            )

        # 원천 데이터 보기 제거(요청사항)

with tabs[1]:
    st.subheader("S관 수주현황(월별 집계)")
    st.caption("수주현황 시트의 품명을 S관 제품 마스터(제품명)와 매칭한 행만 집계합니다.")

    if not order_sheet:
        st.warning("수주현황 시트를 선택할 수 없습니다. '업데이트 데이터/생산실적현황(간편)_S관.xlsx'에 '수주현황' 시트를 저장/업로드한 뒤, 사이드바에서 시트를 선택해주세요.")
        st.stop()

    try:
        order_mtime = Path(order_book_xlsx).stat().st_mtime if Path(order_book_xlsx).exists() else None
        orders_raw = load_order_status_raw(order_book_xlsx, sheet_name=order_sheet, cache_bust=order_mtime)
        orders_s = load_order_status_sgwan(
            order_book_xlsx,
            master_xlsx,
            sheet_name=order_sheet,
            cache_bust=order_mtime,
        )

        misc_raw = pd.DataFrame()
        misc_s = pd.DataFrame()
        if misc_sheet:
            misc_raw = load_order_status_raw(
                order_book_xlsx,
                sheet_name=misc_sheet,
                header_row_1based=1,
                cache_bust=order_mtime,
            )
            misc_s = load_order_status_sgwan(
                order_book_xlsx,
                master_xlsx,
                sheet_name=misc_sheet,
                header_row_1based=1,
                cache_bust=order_mtime,
            )
            misc_s = _normalize_misc_orders_for_dashboard(misc_s)
    except Exception as e:
        st.error(f"수주현황 로드 실패: {e}")
        orders_raw = pd.DataFrame()
        orders_s = pd.DataFrame()
        misc_raw = pd.DataFrame()
        misc_s = pd.DataFrame()

    orders_raw_all = pd.concat([orders_raw, misc_raw], ignore_index=True) if not misc_raw.empty else orders_raw
    orders_s_all = pd.concat([orders_s, misc_s], ignore_index=True, sort=False) if not misc_s.empty else orders_s

    if orders_s_all.empty:
        # If raw exists but filtered is empty, most likely sheet/headers mismatch or master matching fails.
        msg = "S관 수주현황 데이터가 없습니다(매칭 결과 0행이거나 파일/시트/컬럼을 확인하세요)."
        try:
            p = Path(order_book_xlsx)
            if p.exists():
                xl = pd.ExcelFile(p)
                if order_sheet not in xl.sheet_names:
                    msg = f"'{order_sheet}' 시트가 없습니다. 현재 시트: {xl.sheet_names}"
        except Exception:
            pass
        st.info(msg)
    else:
        if not orders_raw_all.empty:
            st.caption(f"S관 제품 매칭 행수: {len(orders_s_all):,} / 전체 행수: {len(orders_raw_all):,}")
        df = orders_s_all.copy()

        # Parse month
        if "__month_date__" in df.columns:
            m = pd.to_datetime(df["__month_date__"], errors="coerce")
            df["월"] = m.dt.to_period("M").dt.strftime("%Y-%m")
            if "연도" not in df.columns:
                df["연도"] = m.dt.year
        else:
            df["월"] = df.get("월", pd.Series([""] * len(df))).astype("string")

        # Normalize numeric columns
        for col in ["오더수량", "수주금액", "수주금액(원)", "수주금액(달러)", "포장 진도율"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        # Normalize date columns for view
        for col in ["수주 전송일", "영업출고요청일", "영업협의출고일", "포장완료일"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce").dt.date

        # Pre-compute aggregations (for consistent layout sizing)
        monthly2 = pd.DataFrame()
        cat2: pd.DataFrame | None = None

        group_cols = ["월"]
        if "구분" in df.columns:
            group_cols.append("구분")
        monthly_agg = {}
        if "오더수량" in df.columns:
            monthly_agg["오더수량 합계"] = ("오더수량", "sum")
        if "수주금액(원)" in df.columns:
            monthly_agg["수주금액(원) 합계"] = ("수주금액(원)", "sum")
        if "수주금액(달러)" in df.columns:
            monthly_agg["수주금액(달러) 합계"] = ("수주금액(달러)", "sum")
        if monthly_agg:
            monthly = df.groupby(group_cols, dropna=False).agg(**monthly_agg).reset_index()
            monthly = monthly.sort_values(group_cols).reset_index(drop=True)
            total_row: dict[str, object] = {c: "" for c in group_cols}
            total_row["월"] = "합계"
            for c in ["오더수량 합계", "수주금액(원) 합계", "수주금액(달러) 합계"]:
                if c in monthly.columns:
                    total_row[c] = float(monthly[c].sum())
            monthly2 = pd.concat([monthly, pd.DataFrame([total_row])], ignore_index=True)

        if "분류요약" in df.columns:
            group_cols = ["분류요약"]
            cat_agg = {}
            if "오더수량" in df.columns:
                cat_agg["오더수량 합계"] = ("오더수량", "sum")
            if "수주금액(원)" in df.columns:
                cat_agg["수주금액(원) 합계"] = ("수주금액(원)", "sum")
            if "수주금액(달러)" in df.columns:
                cat_agg["수주금액(달러) 합계"] = ("수주금액(달러)", "sum")
            if cat_agg:
                cat = df.groupby(group_cols, dropna=False).agg(**cat_agg).reset_index()
                cat = cat.sort_values(group_cols).reset_index(drop=True)
                total_row = {"분류요약": "합계"}
                for c in ["오더수량 합계", "수주금액(원) 합계", "수주금액(달러) 합계"]:
                    if c in cat.columns:
                        total_row[c] = float(cat[c].sum())
                cat2 = pd.concat([cat, pd.DataFrame([total_row])], ignore_index=True)

        left, right = st.columns(2)

        # ===== 월별 집계 =====
        with left:
            with st.container(border=True):
                st.markdown("**월별 집계**")

                group_cols = ["월"]
                if "구분" in monthly2.columns:
                    group_cols.append("구분")
                fmt = {c: "{:,.0f}" for c in monthly2.columns if c not in group_cols}
                st.dataframe(
                    monthly2.style.format(fmt),
                    use_container_width=True,
                    hide_index=True,
                )

                # 월별 집계 박스 높이를 분류요약별 집계 박스에 가깝게 맞춤
                if cat2 is not None and not monthly2.empty:
                    row_px = 34
                    pad_rows = max(0, len(cat2) - len(monthly2))
                    pad_px = min(700, pad_rows * row_px)
                    if pad_px > 0:
                        st.markdown(f"<div style='height:{pad_px}px'></div>", unsafe_allow_html=True)

        # ===== 분류요약별 집계 =====
        with right:
            with st.container(border=True):
                st.markdown("**분류요약별 집계**")
                if "분류요약" not in df.columns:
                    st.info("S관 제품 마스터의 AY열(분류요약) 매핑을 확인하세요. 현재 데이터에 '분류요약' 컬럼이 없습니다.")
                else:
                    if cat2 is None:
                        st.info("분류요약별 집계 데이터가 없습니다.")
                    else:
                        group_cols = ["분류요약"]
                        fmt = {c: "{:,.0f}" for c in cat2.columns if c not in group_cols}
                        st.dataframe(
                            cat2.style.format(fmt),
                            use_container_width=True,
                            hide_index=True,
                        )

        # ===== 원천(필터) =====
        with st.container(border=True):
            st.markdown("**상세 수주 내역**")

        df_view = df.copy()

        show_cols = [
            c
            for c in [
                "월",
                "분류요약",
                "구분",
                "작지번호",
                "고객",
                "품명",
                "담당자",
                "국가",
                "오더수량",
                "수주금액",
                "화폐",
                "수주금액(원)",
                "수주금액(달러)",
                "수주 전송일",
                "영업출고요청일",
                "현재상태",
                "포장 진도율",
                "포장완료일",
            ]
            if c in df_view.columns
        ]
        df_view = df_view.sort_values(["월", "구분", "작지번호"], na_position="last") if all(
            c in df_view.columns for c in ["월", "구분", "작지번호"]
        ) else df_view

        fmt = {}
        for col in ["오더수량", "수주금액", "수주금액(원)", "수주금액(달러)"]:
            if col in df_view.columns:
                fmt[col] = "{:,.0f}"
        if "포장 진도율" in df_view.columns:
            fmt["포장 진도율"] = "{:,.1f}"

        st.dataframe(
            df_view[show_cols].style.format(fmt),
            use_container_width=True,
            hide_index=True,
        )

with tabs[2]:
    st.subheader("APS 리스크관리")
    st.caption("원복 엑셀 1개(`APS 변동사항 체크.xlsx`)로 변동 감지 → 원인 분류 → 리스크/액션 리스트까지 생성합니다.")

    st.markdown("**업데이트할 파일 위치**")
    st.code(str(UPDATE_DATA_DIR), language=None)

    aps_default = _pick_input_path("APS 변동사항 체크.xlsx")
    aps_xlsx = st.text_input("APS 변동사항 체크.xlsx", value=str(aps_default))

    aps_path = Path(aps_xlsx)
    if not aps_path.exists():
        st.warning(f"파일을 찾지 못했습니다: {aps_path}")
        st.info("`업데이트 데이터` 폴더에 `APS 변동사항 체크.xlsx`를 넣어주세요.")
    else:
        uploaded_at = _excel_upload_time_kst(aps_path)
        if uploaded_at:
            st.caption(f"APS 엑셀 수정/업로드 시간: {uploaded_at} (한국시간)")

        scope = st.radio("스코프(공정)", ["기본(총합계+[85]포장)", "전체"], index=0, horizontal=True)
        if scope.startswith("기본"):
            scope_processes = ("총합계", "[85]포장")
            scope_key = "default"
        else:
            xl = pd.ExcelFile(aps_path)
            date_sheets = [s for s in xl.sheet_names if s != "제품명등록" and str(s).isdigit()]
            if not date_sheets:
                st.error("날짜 시트를 찾지 못했습니다(예: 260427 또는 20260427).")
                scope_processes = ("총합계", "[85]포장")
                scope_key = "default"
            else:
                sample = pd.read_excel(aps_path, sheet_name=date_sheets[-1], header=[0, 1], nrows=1)
                procs = [c[0] for c in sample.columns if c[0] != "공정 코드:"]
                scope_processes = tuple(dict.fromkeys(procs).keys())
                scope_key = "all"

        cache_paths = aps_default_cache_paths("data")
        sig = aps_signature(aps_path)

        tables = aps_load_cached_tables(cache_paths, sig=sig, scope_key=scope_key)
        stale_tables = None if tables is not None else aps_load_any_tables(cache_paths)

        lock_path = _aps_lock_path("data")
        status_path = _aps_status_path("data")
        status_obj = _read_json_if_exists(status_path) or {}
        is_running = lock_path.exists() or status_obj.get("state") == "running"

        if tables is None:
            # 자동 백그라운드 분석 시작(중복 방지)
            if not is_running:
                _launch_aps_cache_build(input_path=aps_path, scope=scope_key, days=8)
                is_running = True
                st.info("APS 파일 변경을 감지해 백그라운드 분석을 시작했습니다. (이 탭은 이전 캐시로 먼저 표시됩니다)")
            else:
                st.info("백그라운드 분석이 진행 중입니다. 캐시 생성이 끝나면 자동으로 최신 결과가 표시됩니다.")

            if status_obj:
                st.caption(f"상태: {status_obj.get('state','-')} | 시작: {status_obj.get('started_at','-')} | 종료: {status_obj.get('finished_at','-')}")
                if status_obj.get("state") == "error":
                    st.warning(f"분석 실패: {status_obj.get('error')}")

            if st.button("새로고침"):
                st.rerun()

            tables = stale_tables

        if not tables:
            st.stop()

        if tables is not None:
            st.caption(f"캐시 파일: {cache_paths.data_pkl}")

        if not tables:
            st.stop()

        action_df = tables.get("액션리스트", pd.DataFrame())
        risk_df = tables.get("리스크요약", pd.DataFrame())
        total_df = tables.get("변동분석_총합계", pd.DataFrame())
        pack_df = tables.get("변동분석_포장", pd.DataFrame())

        # 최신 기준일자(=금일 관리 기준)
        try:
            latest_asof = pd.to_datetime(total_df.get("기준일자"), errors="coerce").dt.date.max() if not total_df.empty else None
        except Exception:
            latest_asof = None
        if latest_asof:
            st.caption(f"APS 최신 기준일자(관리 기준): {latest_asof}")

        # 7일 트렌드/금일 액션 요약
        k1, k2, k3 = st.columns(3)
        with k1:
            st.metric("금일 액션 건수", f"{len(action_df):,}" if not action_df.empty else "0")
        with k2:
            if not risk_df.empty and "변동_횟수" in risk_df.columns:
                risky = int((pd.to_numeric(risk_df["변동_횟수"], errors="coerce").fillna(0) > 0).sum())
                st.metric("7일 변동 수주", f"{risky:,}")
            else:
                st.metric("7일 변동 수주", "-")
        with k3:
            if not risk_df.empty and "최대_지연일수" in risk_df.columns:
                mx = pd.to_numeric(risk_df["최대_지연일수"], errors="coerce").max()
                st.metric("7일 최대 지연(일)", f"{int(mx):,}" if pd.notna(mx) else "-")
            else:
                st.metric("7일 최대 지연(일)", "-")

        if not risk_df.empty and "윈도우_시작" in risk_df.columns and "윈도우_종료" in risk_df.columns:
            try:
                w_start = pd.to_datetime(risk_df["윈도우_시작"], errors="coerce").dt.date.min()
                w_end = pd.to_datetime(risk_df["윈도우_종료"], errors="coerce").dt.date.max()
                if w_start and w_end:
                    st.caption(f"7일 트렌드 윈도우: {w_start} ~ {w_end}")
            except Exception:
                pass

        # 수주현황(영업출고요청일/고객/작지번호) 붙이기
        orders_for_join = pd.DataFrame()
        try:
            if order_sheet and Path(order_book_xlsx).exists() and Path(master_xlsx).exists():
                order_mtime = Path(order_book_xlsx).stat().st_mtime
                orders_for_join = load_orders_sgwan_for_join(
                    order_book_xlsx,
                    master_xlsx,
                    order_sheet=order_sheet,
                    misc_sheet=misc_sheet,
                    cache_bust=order_mtime,
                )
        except Exception:
            orders_for_join = pd.DataFrame()

        if not action_df.empty and not orders_for_join.empty and "제품명코드" in action_df.columns:
            o = orders_for_join.copy()
            o = o[o.get("제품명코드").notna()].copy() if "제품명코드" in o.columns else pd.DataFrame()

            # (제품명코드, 요청기준일) 단위로 중복 집계
            if not o.empty and "요청기준일" in o.columns:
                code_counts = o.groupby("제품명코드", dropna=False).size().to_dict() if "제품명코드" in o.columns else {}
                agg = {
                    "작지번호": lambda s: ", ".join([x for x in pd.Series(s).dropna().astype(str).unique().tolist()][:5]),
                    "고객": lambda s: ", ".join([x for x in pd.Series(s).dropna().astype(str).unique().tolist()][:3]),
                }
                if "오더수량" in o.columns:
                    agg["오더수량"] = "sum"
                if "수주금액(원)" in o.columns:
                    agg["수주금액(원)"] = "sum"

                o2 = (
                    o.groupby(["제품명코드", "요청기준일"], dropna=False)
                    .agg(agg)
                    .reset_index()
                    .rename(
                        columns={
                            "작지번호": "수주현황_작지번호",
                            "고객": "수주현황_고객",
                            "오더수량": "수주현황_오더수량",
                            "수주금액(원)": "수주현황_수주금액(원)",
                            "요청기준일": "수주현황_요청기준일",
                        }
                    )
                )

                a = action_df.copy()
                if "고객납기일" in a.columns:
                    a["고객납기일"] = pd.to_datetime(a["고객납기일"], errors="coerce").dt.date
                a["__고객납기_dt__"] = pd.to_datetime(a.get("고객납기일"), errors="coerce")
                o2["__요청일_dt__"] = pd.to_datetime(o2["수주현황_요청기준일"], errors="coerce")

                # 1) exact match (제품명코드 + 고객납기일=영업출고요청일)
                exact = a.merge(
                    o2,
                    left_on=["제품명코드", "__고객납기_dt__"],
                    right_on=["제품명코드", "__요청일_dt__"],
                    how="left",
                )
                exact["수주현황_매칭"] = np.where(exact["수주현황_고객"].notna(), "exact", "")

                # 2) nearest fallback (제품명코드 기준, ±30일)
                def _nearest_fill(ex: pd.DataFrame, candidates: pd.DataFrame, *, tol_days: int) -> None:
                    if ex.empty:
                        return
                    if candidates.empty or "__요청일_dt__" not in candidates.columns:
                        return
                    cand_by: dict[str, pd.DataFrame] = {}
                    for code, g in candidates.groupby("제품명코드", dropna=False):
                        gs = g.sort_values("__요청일_dt__").reset_index(drop=True)
                        cand_by[str(code)] = gs

                    miss_mask = ex["수주현황_고객"].isna()
                    if not miss_mask.any():
                        return

                    def _one(i: int) -> None:
                        r = ex.loc[i]
                        code = str(r.get("제품명코드", ""))
                        tgt = r.get("__고객납기_dt__")
                        if not code or pd.isna(tgt) or code not in cand_by:
                            return
                        g = cand_by[code]
                        dates = pd.to_datetime(g["__요청일_dt__"], errors="coerce")
                        dates = dates.dropna().sort_values().to_numpy()
                        if len(dates) == 0:
                            return
                        tgt_ts = pd.to_datetime(tgt, errors="coerce")
                        if pd.isna(tgt_ts):
                            return
                        pos = int(np.searchsorted(dates, tgt_ts.to_datetime64()))
                        cand = []
                        if 0 <= pos < len(dates):
                            cand.append(dates[pos])
                        if 0 <= pos - 1 < len(dates):
                            cand.append(dates[pos - 1])
                        if not cand:
                            return
                        best = min(cand, key=lambda x: abs((x - tgt_ts).days))
                        delta = int((best - tgt_ts).days)
                        if abs(delta) > int(tol_days):
                            return

                        row = g[pd.to_datetime(g["__요청일_dt__"], errors="coerce") == best].head(1)
                        if row.empty:
                            return
                        row0 = row.iloc[0]
                        for col in ["영업출고요청일", "수주현황_작지번호", "수주현황_고객", "수주현황_오더수량", "수주현황_수주금액(원)"]:
                            if col in ex.columns and col in row0.index:
                                ex.at[i, col] = row0[col]
                        ex.at[i, "수주현황_매칭"] = "nearest"

                    for idx in ex.index[miss_mask].tolist():
                        _one(idx)

                _nearest_fill(exact, o2, tol_days=30)

                # 3) 디버그: ±30일에 못 맞춘 건들에 대해, ±365일 기준 최근접 요청일/차이 저장
                try:
                    # tolerance 없이 '가장 가까운' 요청일을 계산(최대 ±365일)
                    miss_any = exact["수주현황_고객"].isna()
                    if miss_any.any() and not o2.empty:
                        for i in exact.index[miss_any].tolist():
                            code = str(exact.at[i, "제품명코드"]) if "제품명코드" in exact.columns else ""
                            tgt = exact.at[i, "__고객납기_dt__"] if "__고객납기_dt__" in exact.columns else pd.NaT
                            if not code or pd.isna(tgt):
                                continue
                            g = o2[o2["제품명코드"].astype(str) == code].copy()
                            if g.empty:
                                continue
                            dates = pd.to_datetime(g["__요청일_dt__"], errors="coerce").dropna()
                            if dates.empty:
                                continue
                            tgt_ts = pd.to_datetime(tgt, errors="coerce")
                            if pd.isna(tgt_ts):
                                continue
                            best = dates.iloc[(dates - tgt_ts).abs().argsort()[:1]].iloc[0]
                            delta = int((best - tgt_ts).days)
                            if abs(delta) <= 365:
                                exact.at[i, "수주현황_최근접요청일"] = best.date()
                                exact.at[i, "수주현황_요청일차이(일)"] = delta
                except Exception:
                    pass

                if "변경 납기일(당일 종료예정일)" in exact.columns and "영업출고요청일" in exact.columns:
                    exact["초과일(변경-요청)"] = (
                        pd.to_datetime(exact["변경 납기일(당일 종료예정일)"], errors="coerce")
                        - pd.to_datetime(exact.get("수주현황_요청기준일"), errors="coerce")
                    ).dt.days

                drop_cols = [c for c in ["__고객납기_dt__", "__요청일_dt__"] if c in exact.columns]
                action_df = exact.drop(columns=drop_cols, errors="ignore")

        # 수주현황 고객이 None/NaN인 건들 원인 표시(품명 매핑/요청일 매칭 문제)
        if not action_df.empty and "수주현황_고객" in action_df.columns:
            miss = action_df["수주현황_고객"].isna()
            if miss.any():
                def _reason(r: pd.Series) -> str:
                    code = r.get("제품명코드")
                    if code is None or str(code).strip() == "" or str(code).lower() == "nan":
                        return "제품명코드 없음"
                    if "code_counts" in locals():
                        code_set = {str(k) for k in code_counts.keys()}
                        if str(code) not in code_set:
                            return "수주현황에 제품 없음(품명/마스터 매핑)"
                    near_dt = r.get("수주현황_최근접요청일")
                    delta = r.get("수주현황_요청일차이(일)")
                    if pd.notna(near_dt) and pd.notna(delta):
                        try:
                            return f"요청일이 멀어요({int(delta):+d}일) / 최근접:{near_dt}"
                        except Exception:
                            pass
                    return "요청일 매칭 실패(±30일 내 없음)"

                action_df = action_df.copy()
                action_df.loc[miss, "수주현황_미매칭사유"] = action_df[miss].apply(_reason, axis=1)
                action_df.loc[~miss, "수주현황_미매칭사유"] = ""

        t1, t2, t3, t4 = st.tabs(["액션리스트", "리스크요약(7일)", "변동분석_총합계", "변동분석_포장"])
        with t1:
            st.caption("업무 처리용(최신 기준일자, 총합계 이벤트)")
            view = action_df.copy()
            if "수주현황_요청기준일" in view.columns:
                view["요청납기일"] = pd.to_datetime(view["수주현황_요청기준일"], errors="coerce").dt.date
            if "고객납기일" in view.columns:
                # 수주현황 매칭 실패(영업출고요청일 없음)인 경우 APS 고객납기일로 보완
                view["요청납기일"] = pd.to_datetime(view.get("요청납기일"), errors="coerce").dt.date
                view["요청납기일"] = view["요청납기일"].fillna(pd.to_datetime(view["고객납기일"], errors="coerce").dt.date)

            # 표준 표시 컬럼(간소화)
            show_cols = [
                "제품구분",
                "이니셜",
                "수주번호",
                "제품명코드",
                "제품명_마스터",
                "수주현황_고객",
                "수주현황_미매칭사유",
                "수주현황_최근접요청일",
                "수주현황_요청일차이(일)",
                "요청납기일",
                "변경 납기일(당일 종료예정일)",
                "기존 납기일(전일 종료예정일)",
                "변동일수",
                "초과일(변경-요청)",
                "원인",
                "조치유형",
                "SKU수",
            ]
            show_cols = [c for c in show_cols if c in view.columns]
            if "수주현황_고객" in view.columns:
                n_none = int(view["수주현황_고객"].isna().sum())
                if n_none > 0:
                    st.warning(f"수주현황 고객 매칭 실패: {n_none:,}건 (수주현황/마스터 매핑 또는 요청일 매칭 문제)")
            st.dataframe(view[show_cols], use_container_width=True, hide_index=True)
        with t2:
            st.caption("불안정 수주 요약(최근 7일, 총합계 기준)")
            st.dataframe(risk_df, use_container_width=True, hide_index=True)
        with t3:
            st.dataframe(total_df, use_container_width=True, hide_index=True)
        with t4:
            st.dataframe(pack_df, use_container_width=True, hide_index=True)

        out_buf = io.BytesIO()
        with st.expander("결과 다운로드", expanded=False):
            if st.button("다운로드 파일 생성"):
                with pd.ExcelWriter(out_buf, engine="openpyxl") as w:
                    for name, df in tables.items():
                        df.to_excel(w, sheet_name=name, index=False)
                st.download_button(
                    "결과 엑셀 다운로드",
                    data=out_buf.getvalue(),
                    file_name="aps_risk_result.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
