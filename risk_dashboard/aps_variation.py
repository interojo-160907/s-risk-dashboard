from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Literal

import numpy as np
import pandas as pd


META_GROUP = "공정 코드:"


@dataclass(frozen=True)
class APSVariationCols:
    기준일자: str = "기준일자"
    수주번호: str = "수주번호"
    수요제품코드: str = "수요제품코드"
    제품이름: str = "제품 이름"
    납기일: str = "납기일"  # 고객납기(요청)
    제품그룹코드: str = "제품 그룹 코드"
    이니셜: str = "이니셜"

    제품명코드: str = "제품명코드"
    공정코드: str = "공정코드"
    종료예정일: str = "종료예정일"
    필요수량: str = "필요수량"

    제품구분: str = "제품구분"
    제품명_마스터: str = "제품명_마스터"
    거래처명: str = "거래처명"
    공장구분: str = "공장구분"

    전일_종료예정일: str = "전일 종료예정일"
    전일_필요수량: str = "전일 필요수량"
    변동일수: str = "변동일수"
    수량변동: str = "수량변동"
    이벤트: str = "이벤트"

    원인: str = "원인"
    조치유형: str = "조치유형"


COLS = APSVariationCols()


def parse_sheet_date(sheet_name: str) -> date | None:
    s = str(sheet_name).strip()
    if not s.isdigit():
        return None
    if len(s) == 6:  # yymmdd
        yy = int(s[:2])
        mm = int(s[2:4])
        dd = int(s[4:6])
        return date(2000 + yy, mm, dd)
    if len(s) == 8:  # yyyymmdd
        yyyy = int(s[:4])
        mm = int(s[4:6])
        dd = int(s[6:8])
        return date(yyyy, mm, dd)
    return None


def derive_제품명코드(수요제품코드: object) -> str | None:
    if pd.isna(수요제품코드):
        return None
    s = str(수요제품코드).strip()
    if not s:
        return None
    if s.startswith("T") and len(s) >= 5:
        return s[:5]
    if s.startswith("S") and len(s) >= 4:
        return s[:4]
    return None


def _to_date_series(s: pd.Series) -> pd.Series:
    dt = pd.to_datetime(s, errors="coerce")
    return dt.dt.date


def _read_master_products(xlsx_path: str | Path) -> pd.DataFrame:
    master = pd.read_excel(xlsx_path, sheet_name="제품명등록")
    master = master.copy()
    master["제품명코드"] = master["제품명코드"].astype(str).str.strip()
    master = master[master["제품명코드"].ne("")].copy()
    out = master[["제품명코드"]].copy()
    if "공장구분" in master.columns:
        out[COLS.공장구분] = master["공장구분"].astype(str).str.strip()
    # 제품구분(표시용): 생산제품군 > 판매제품군 > 구분
    prod_group_col = "생산제품군" if "생산제품군" in master.columns else None
    if prod_group_col is None and "판매제품군" in master.columns:
        prod_group_col = "판매제품군"
    if prod_group_col is None and "구분" in master.columns:
        prod_group_col = "구분"
    if prod_group_col:
        out[COLS.제품구분] = master[prod_group_col]
    if "제품명" in master.columns:
        out[COLS.제품명_마스터] = master["제품명"]
    if "거래처명" in master.columns:
        out[COLS.거래처명] = master["거래처명"]
    return out.drop_duplicates(subset=["제품명코드"], keep="last")


def _date_sheet_names(xlsx_path: str | Path) -> list[str]:
    xl = pd.ExcelFile(xlsx_path)
    names: list[str] = []
    for s in xl.sheet_names:
        if s == "제품명등록":
            continue
        if parse_sheet_date(s) is not None:
            names.append(s)
    return names


def _filter_summary_rows(df: pd.DataFrame) -> pd.DataFrame:
    key = (META_GROUP, "수주번호")
    if key not in df.columns:
        return df
    s = df[key].astype(str).str.strip()
    # '총합계' 같은 요약 행 제거(원복 엑셀에 섞여 있을 수 있음)
    drop = s.str.contains(r"총합계|합계", na=False)
    return df[~drop].copy()


def _build_long_for_sheet(
    xlsx_path: str | Path,
    sheet_name: str,
    *,
    scope_processes: Iterable[str],
) -> pd.DataFrame:
    기준일자 = parse_sheet_date(sheet_name)
    if 기준일자 is None:
        raise ValueError(f"날짜 시트명이 아닙니다: {sheet_name!r}")

    raise RuntimeError(
        "_build_long_for_sheet는 더 이상 직접 호출되지 않습니다. "
        "build_aps_long_table에서 openpyxl 워크북을 1회만 열어 처리합니다."
    )


def _build_long_for_ws(
    ws: object,
    *,
    sheet_name: str,
    기준일자: date,
    scope_processes: Iterable[str],
) -> pd.DataFrame:
    header_rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if len(header_rows) < 2:
        return pd.DataFrame()

    r1 = list(header_rows[0])
    r2 = list(header_rows[1])
    proc = ""
    tuples: list[tuple[str, str]] = []
    for a, b in zip(r1, r2, strict=False):
        if a is not None and str(a).strip() != "":
            proc = str(a).strip()
        sub = "" if b is None else str(b).strip()
        tuples.append((proc, sub))

    idx_map: dict[tuple[str, str], int] = {t: i for i, t in enumerate(tuples)}

    meta_keys = {
        (META_GROUP, "수주번호"): COLS.수주번호,
        (META_GROUP, "수요 제품 코드"): COLS.수요제품코드,
        (META_GROUP, "제품 이름"): COLS.제품이름,
        (META_GROUP, "납기일"): COLS.납기일,
        (META_GROUP, "제품 그룹 코드"): COLS.제품그룹코드,
        (META_GROUP, "이니셜"): COLS.이니셜,
    }
    missing_meta = [k for k in meta_keys if k not in idx_map]
    if missing_meta:
        raise ValueError(f"{sheet_name}: 메타 컬럼 누락: {missing_meta}")

    meta_idxs = {meta_keys[k]: idx_map[k] for k in meta_keys}
    proc_cols: dict[str, tuple[int | None, int | None]] = {}
    for p in scope_processes:
        qty_idx = idx_map.get((p, "생산 수량"))
        end_idx = idx_map.get((p, "종료일"))
        proc_cols[p] = (qty_idx, end_idx)

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        su = row[meta_idxs[COLS.수주번호]]
        if su is None:
            continue
        su_s = str(su).strip()
        if not su_s or ("합계" in su_s) or ("총합계" in su_s):
            continue

        demand_code = row[meta_idxs[COLS.수요제품코드]]
        prod_name = row[meta_idxs[COLS.제품이름]]
        due = row[meta_idxs[COLS.납기일]]
        grp = row[meta_idxs[COLS.제품그룹코드]]
        ini = row[meta_idxs[COLS.이니셜]]

        for proc_name, (qty_i, end_i) in proc_cols.items():
            if qty_i is None and end_i is None:
                continue
            qty = row[qty_i] if qty_i is not None else None
            end = row[end_i] if end_i is not None else None
            if qty in [None, ""] and end in [None, ""]:
                continue

            rows.append(
                {
                    COLS.기준일자: 기준일자,
                    COLS.수주번호: su_s,
                    COLS.수요제품코드: None if demand_code is None else str(demand_code).strip(),
                    COLS.제품이름: None if prod_name is None else str(prod_name).strip(),
                    COLS.납기일: due,
                    COLS.제품그룹코드: None if grp is None else str(grp).strip(),
                    COLS.이니셜: None if ini is None else str(ini).strip(),
                    COLS.공정코드: proc_name,
                    COLS.필요수량: qty,
                    COLS.종료예정일: end,
                }
            )

    if not rows:
        return pd.DataFrame()

    out = pd.DataFrame(rows)
    out[COLS.납기일] = _to_date_series(out[COLS.납기일])
    out[COLS.종료예정일] = _to_date_series(out[COLS.종료예정일])
    out[COLS.필요수량] = pd.to_numeric(out[COLS.필요수량], errors="coerce")
    return out


def build_aps_long_table(
    xlsx_path: str | Path,
    *,
    scope_processes: Iterable[str] = ("총합계", "[85]포장"),
    window_days: int | None = 8,
) -> pd.DataFrame:
    master = _read_master_products(xlsx_path)
    # S관만 관리(마스터에 공장구분이 있으면 S관만)
    master2 = master.copy()
    if COLS.공장구분 in master2.columns:
        master2["_plant_norm"] = master2[COLS.공장구분].astype(str).str.strip()
        master2 = master2[master2["_plant_norm"].str.contains("S관", na=False)].copy()
        master2 = master2.drop(columns=["_plant_norm"], errors="ignore")

    managed_codes = set(master2["제품명코드"].astype(str).str.strip().tolist())

    # 날짜 시트 파싱 + 성능 최적화: 워크북은 1회만 open
    try:
        import openpyxl  # type: ignore[import-not-found]
    except Exception as e:  # pragma: no cover
        raise RuntimeError("openpyxl이 필요합니다(requirements.txt 확인)") from e

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        pairs2: list[tuple[str, date]] = []
        for s in wb.sheetnames:
            if s == "제품명등록":
                continue
            d = parse_sheet_date(str(s))
            if d is not None:
                pairs2.append((str(s), d))
        pairs2.sort(key=lambda x: x[1])

        if window_days is not None and window_days > 0:
            pairs2 = pairs2[-window_days:]

        longs: list[pd.DataFrame] = []
        for sheet, d in pairs2:
            ws = wb[sheet]
            longs.append(_build_long_for_ws(ws, sheet_name=sheet, 기준일자=d, scope_processes=scope_processes))
    finally:
        wb.close()

    out = pd.concat([x for x in longs if not x.empty], ignore_index=True) if "longs" in locals() else pd.DataFrame()
    if out.empty:
        return out

    out[COLS.제품명코드] = out[COLS.수요제품코드].map(derive_제품명코드)
    out = out[out[COLS.제품명코드].isin(managed_codes)].copy()

    out = out.merge(master2, on=COLS.제품명코드, how="left")
    return out


def add_daily_deltas(df_long: pd.DataFrame) -> pd.DataFrame:
    if df_long.empty:
        return df_long

    out = df_long.copy()

    key = [COLS.수주번호, COLS.수요제품코드, COLS.공정코드]
    out[COLS.기준일자] = pd.to_datetime(out[COLS.기준일자], errors="coerce").dt.date
    out[COLS.종료예정일] = pd.to_datetime(out[COLS.종료예정일], errors="coerce").dt.date

    out = out.sort_values(key + [COLS.기준일자]).copy()
    out[COLS.전일_종료예정일] = out.groupby(key, dropna=False)[COLS.종료예정일].shift(1)
    out[COLS.전일_필요수량] = out.groupby(key, dropna=False)[COLS.필요수량].shift(1)

    out[COLS.변동일수] = (
        pd.to_datetime(out[COLS.종료예정일], errors="coerce")
        - pd.to_datetime(out[COLS.전일_종료예정일], errors="coerce")
    ).dt.days
    out[COLS.수량변동] = pd.to_numeric(out[COLS.필요수량], errors="coerce") - pd.to_numeric(
        out[COLS.전일_필요수량], errors="coerce"
    )

    out[COLS.이벤트] = (out[COLS.변동일수].fillna(0) > 0) | (out[COLS.수량변동].fillna(0) > 0)
    return out


Cause = Literal["수주 증가", "생산 부족", "포장 영향", "APS 재계산", ""]


def classify_cause(df_deltas: pd.DataFrame) -> pd.DataFrame:
    if df_deltas.empty:
        return df_deltas

    out = df_deltas.copy()
    out[COLS.원인] = ""

    qty_inc = out[COLS.수량변동].fillna(0) > 0
    delay = out[COLS.변동일수].fillna(0) > 0
    is_total = out[COLS.공정코드] == "총합계"

    # 총합계 행에서 '포장 영향' 판단을 위해 동일 기준일자/수주/제품의 포장 지연을 가져옴
    pack = out[out[COLS.공정코드] == "[85]포장"][
        [COLS.기준일자, COLS.수주번호, COLS.수요제품코드, COLS.변동일수]
    ].rename(columns={COLS.변동일수: "포장_변동일수"})
    out = out.merge(pack, on=[COLS.기준일자, COLS.수주번호, COLS.수요제품코드], how="left")
    pack_delay = out["포장_변동일수"].fillna(0) > 0

    # 우선순위: 수주 증가 > (총합계) 포장 영향 > 생산 부족 > APS 재계산
    out.loc[qty_inc, COLS.원인] = "수주 증가"
    out.loc[~qty_inc & is_total & delay & pack_delay, COLS.원인] = "포장 영향"
    out.loc[~qty_inc & delay & (out[COLS.원인] == ""), COLS.원인] = "생산 부족"
    out.loc[(out[COLS.이벤트].fillna(False)) & (out[COLS.원인] == ""), COLS.원인] = "APS 재계산"

    out = out.drop(columns=["포장_변동일수"], errors="ignore")
    return out


def build_risk_summary(df: pd.DataFrame, *, days: int = 7) -> pd.DataFrame:
    if df.empty:
        return df

    base = df[df[COLS.공정코드] == "총합계"].copy()
    if base.empty:
        return base

    max_day = pd.to_datetime(base[COLS.기준일자], errors="coerce").dt.date.max()
    if max_day is None or pd.isna(max_day):
        return pd.DataFrame()
    start_day = max_day.fromordinal(max_day.toordinal() - (days - 1))
    win = base[(base[COLS.기준일자] >= start_day) & (base[COLS.기준일자] <= max_day)].copy()

    def _mode(series: pd.Series) -> str:
        s = series.dropna().astype(str)
        s = s[s.ne("")]
        if s.empty:
            return ""
        return s.mode().iloc[0]

    # 수주번호+제품명코드 단위(해당 제품군 내 SKU 중 가장 늦는 종료예정일이 실 납기라고 가정)
    per_day = (
        win.groupby([COLS.기준일자, COLS.수주번호, COLS.제품명코드], dropna=False)
        .agg(
            종료예정일_max=(COLS.종료예정일, lambda x: pd.to_datetime(pd.Series(x), errors="coerce").max()),
            이벤트_any=(COLS.이벤트, lambda x: bool(pd.Series(x).fillna(False).any())),
            대표_원인=(COLS.원인, _mode),
            고객납기일=(COLS.납기일, lambda x: pd.to_datetime(pd.Series(x), errors="coerce").dt.date.mode().iloc[0] if pd.to_datetime(pd.Series(x), errors="coerce").notna().any() else None),
            제품구분=(COLS.제품구분, _mode),
            이니셜=(COLS.이니셜, _mode),
            **{COLS.제품그룹코드: (COLS.제품그룹코드, _mode)},
            제품명_마스터=(COLS.제품명_마스터, _mode),
            거래처명=(COLS.거래처명, _mode),
        )
        .reset_index()
    )

    def _max_delay_days(series: pd.Series) -> float:
        s = pd.to_datetime(series, errors="coerce")
        if s.notna().any():
            return float((s.max() - s.min()).days)
        return float("nan")

    agg = (
        per_day.groupby([COLS.수주번호, COLS.제품명코드], dropna=False)
        .agg(
            변동_횟수=("이벤트_any", lambda x: int(pd.Series(x).fillna(False).sum())),
            최대_지연일수=("종료예정일_max", _max_delay_days),
            대표_원인=("대표_원인", _mode),
            고객납기일=("고객납기일", _mode),
            제품구분=("제품구분", _mode),
            이니셜=("이니셜", _mode),
            **{COLS.제품그룹코드: (COLS.제품그룹코드, _mode)},
            제품명_마스터=("제품명_마스터", _mode),
            거래처명=("거래처명", _mode),
        )
        .reset_index()
    )
    agg["윈도우_시작"] = start_day
    agg["윈도우_종료"] = max_day
    return agg.sort_values(["변동_횟수", "최대_지연일수"], ascending=[False, False])


def build_action_list(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    max_day = pd.to_datetime(df[COLS.기준일자], errors="coerce").dt.date.max()
    if max_day is None or pd.isna(max_day):
        return pd.DataFrame()

    base = df[(df[COLS.기준일자] == max_day) & (df[COLS.공정코드] == "총합계") & (df[COLS.이벤트])].copy()
    if base.empty:
        return base

    def _mode(series: pd.Series) -> str:
        s = series.dropna().astype(str)
        s = s[s.ne("")]
        if s.empty:
            return ""
        return s.mode().iloc[0]

    def _pick_cause(series: pd.Series) -> str:
        priority = ["수주 증가", "포장 영향", "생산 부족", "APS 재계산"]
        vals = [v for v in series.dropna().astype(str).tolist() if v]
        for p in priority:
            if p in vals:
                return p
        return _mode(series)

    grouped = (
        base.groupby([COLS.수주번호, COLS.제품명코드], dropna=False)
        .agg(
            제품구분=(COLS.제품구분, _mode),
            이니셜=(COLS.이니셜, _mode),
            **{COLS.제품그룹코드: (COLS.제품그룹코드, _mode)},
            고객납기일=(COLS.납기일, lambda x: pd.to_datetime(pd.Series(x), errors="coerce").dt.date.mode().iloc[0] if pd.to_datetime(pd.Series(x), errors="coerce").notna().any() else None),
            기존_종료예정일=(COLS.전일_종료예정일, lambda x: pd.to_datetime(pd.Series(x), errors="coerce").max().date() if pd.to_datetime(pd.Series(x), errors="coerce").notna().any() else None),
            변경_종료예정일=(COLS.종료예정일, lambda x: pd.to_datetime(pd.Series(x), errors="coerce").max().date() if pd.to_datetime(pd.Series(x), errors="coerce").notna().any() else None),
            수량변동=(COLS.수량변동, lambda x: float(pd.to_numeric(pd.Series(x), errors="coerce").fillna(0).sum())),
            원인=(COLS.원인, _pick_cause),
            제품명_마스터=(COLS.제품명_마스터, _mode),
            거래처명=(COLS.거래처명, _mode),
            SKU수=(COLS.수요제품코드, lambda x: int(pd.Series(x).dropna().astype(str).nunique())),
        )
        .reset_index()
    )

    grouped["기존 납기일(전일 종료예정일)"] = grouped["기존_종료예정일"]
    grouped["변경 납기일(당일 종료예정일)"] = grouped["변경_종료예정일"]
    grouped[COLS.변동일수] = (
        pd.to_datetime(grouped["변경_종료예정일"], errors="coerce") - pd.to_datetime(grouped["기존_종료예정일"], errors="coerce")
    ).dt.days
    grouped["초과일(변경-고객)"] = (
        pd.to_datetime(grouped["변경_종료예정일"], errors="coerce") - pd.to_datetime(grouped["고객납기일"], errors="coerce")
    ).dt.days

    action_map = {
        "수주 증가": "납기 협의",
        "생산 부족": "생산 우선순위 조정",
        "포장 영향": "포장 협의",
        "APS 재계산": "내부 검토",
        "": "",
    }
    grouped[COLS.조치유형] = grouped[COLS.원인].map(action_map).fillna("")
    grouped["협의상태"] = ""
    grouped["비고"] = ""

    cols = [
        COLS.제품구분,
        COLS.이니셜,
        COLS.제품그룹코드,
        COLS.수주번호,
        COLS.제품명코드,
        "고객납기일",
        "기존 납기일(전일 종료예정일)",
        "변경 납기일(당일 종료예정일)",
        COLS.변동일수,
        "초과일(변경-고객)",
        COLS.원인,
        COLS.조치유형,
        "SKU수",
        COLS.수량변동,
        "협의상태",
        "비고",
        COLS.제품명_마스터,
        COLS.거래처명,
    ]
    cols = [c for c in cols if c in grouped.columns]
    return grouped[cols].sort_values([COLS.변동일수, "초과일(변경-고객)"], ascending=[False, False])


def analyze_workbook(
    xlsx_path: str | Path,
    *,
    scope_processes: Iterable[str] = ("총합계", "[85]포장"),
    window_days: int | None = 8,
) -> dict[str, pd.DataFrame]:
    long_df = build_aps_long_table(xlsx_path, scope_processes=scope_processes, window_days=window_days)
    deltas = add_daily_deltas(long_df)
    classified = classify_cause(deltas)

    # 공정별 변동분석 시트(현행 산출 포맷에 맞춤)
    total = classified[classified[COLS.공정코드] == "총합계"].copy()
    pack = classified[classified[COLS.공정코드] == "[85]포장"].copy()

    risk = build_risk_summary(classified, days=7)
    actions = build_action_list(classified)

    return {
        "변동분석_총합계": total,
        "변동분석_포장": pack,
        "리스크요약": risk,
        "액션리스트": actions,
    }


def write_analysis_to_excel(
    out_path: str | Path,
    tables: dict[str, pd.DataFrame],
    *,
    xlsx_path: str | Path,
    scope_processes: Iterable[str],
) -> Path:
    out_p = Path(out_path)
    out_p.parent.mkdir(parents=True, exist_ok=True)

    info_rows = [
        ("입력파일", str(Path(xlsx_path).resolve())),
        ("생성시각", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("스코프공정", ", ".join(list(scope_processes))),
    ]
    info = pd.DataFrame(info_rows, columns=["키", "값"])

    with pd.ExcelWriter(out_p, engine="openpyxl") as w:
        for name, df in tables.items():
            (df if df is not None else pd.DataFrame()).to_excel(w, sheet_name=name, index=False)
        info.to_excel(w, sheet_name="정보", index=False)

    return out_p
